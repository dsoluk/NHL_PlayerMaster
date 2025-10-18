

import csv
import requests
import json
import os
import unicodedata
import jellyfish

# ---- Team and Position Normalization Helpers ----

def load_team_mappings(xlsx_path='Team2TM.xlsx'):
    """Load team code mappings from Team2TM.xlsx.
    Primary requirement: read the NHL column to get the list of valid NHL team abbreviations.
    If HR, NST, or AG columns are present, also build cross-source maps to NHL, but they are optional.
    Returns a dict with sub-maps: { 'HR': {hr: nhl}, 'NST': {nst: nhl}, 'AG': {ag: nhl}, 'NHL': set([...]) }.
    If the Excel file or required column is unavailable, returns an empty/default mapping.
    """
    mappings = {'HR': {}, 'NST': {}, 'AG': {}, 'NHL': set()}

    try:
        import openpyxl  # type: ignore
        xlsx_try_paths = [xlsx_path, os.path.join(os.getcwd(), xlsx_path)]
        xlsx_path_use = next((p for p in xlsx_try_paths if p and os.path.exists(p)), None)
        if not xlsx_path_use:
            print("Team mapping notice: Excel file Team2TM.xlsx not found; proceeding without cross-source mapping.")
            return mappings
        wb = openpyxl.load_workbook(xlsx_path_use, data_only=True)
        ws = wb.active
        # Detect header row
        headers = {}
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        for i, cell in enumerate(first_row):
            key = (str(cell) if cell is not None else '').strip().upper()
            headers[key] = i
        nhl_idx = headers.get('NHL')
        hr_idx = headers.get('HR')
        nst_idx = headers.get('NST')
        ag_idx = headers.get('AG')
        if nhl_idx is None:
            print("Team mapping: 'NHL' column not found in Excel header; proceeding without mapping from Excel.")
            return mappings
        for row in ws.iter_rows(min_row=2, values_only=True):
            nhl = (str(row[nhl_idx]) if row and row[nhl_idx] is not None else '').strip().upper()
            if not nhl:
                continue
            mappings['NHL'].add(nhl)
            if hr_idx is not None and row and row[hr_idx] is not None:
                hr = str(row[hr_idx]).strip().upper()
                if hr:
                    mappings['HR'][hr] = nhl
            if nst_idx is not None and row and row[nst_idx] is not None:
                nst = str(row[nst_idx]).strip().upper()
                if nst:
                    mappings['NST'][nst] = nhl
            if ag_idx is not None and row and row[ag_idx] is not None:
                ag = str(row[ag_idx]).strip().upper()
                if ag:
                    mappings['AG'][ag] = nhl
        return mappings
    except ModuleNotFoundError:
        print("Team mapping notice: openpyxl not installed; proceeding without cross-source mapping.")
    except Exception as e:
        print(f"Team mapping Excel load warning: {e}. Proceeding without cross-source mapping.")

    return mappings


def normalize_position(pos):
    """Map raw position codes to G, D, or F (all others become F)."""
    p = (pos or '').strip().upper()
    if p == 'G':
        return 'G'
    if p == 'D':
        return 'D'
    return 'F' if p else ''


def normalize_team_for_source(source, team_abbrev, mappings):
    """Normalize a team code from a given source to NHL code using mappings.
    Order:
    1) Clean obvious artifacts (BOM, whitespace, casing)
    2) Use source-specific mapping (HR/NST -> NHL) when available
    3) If still not NHL-style, apply common alias fixes (e.g., 'T.B' -> 'TBL')
    4) As a final fallback, strip punctuation (like dots) and map well-known 2-letter forms
    """
    src = (source or '').strip().lower()
    raw = (team_abbrev or '')
    t = raw.replace('\ufeff', '').strip().upper()
    if not t:
        return t

    # First, try explicit source mappings (preferred)
    if src in ('hockey_reference', 'hr', 'hockey-reference'):
        mapped = mappings.get('HR', {}).get(t)
        if mapped:
            return mapped
    elif src in ('naturalstattrick', 'nst'):
        mapped = mappings.get('NST', {}).get(t)
        if mapped:
            return mapped
    elif src in ('ag', 'ag_projections'):
        mapped = mappings.get('AG', {}).get(t)
        if mapped:
            return mapped

    # Common dotted/alias forms seen across sites
    alias_map = {
        'T.B': 'TBL',
        'L.A': 'LAK',
        'N.J': 'NJD',
        'S.J': 'SJS',
        'M.T.L': 'MTL',
        'N.Y.I': 'NYI',
        'N.Y.R': 'NYR',
        'VEG': 'VGK',
    }
    if t in alias_map:
        return alias_map[t]

    # If we have a CSV/Excel list of valid NHL codes, return t if it is one of them
    nhl_set = mappings.get('NHL') or set()
    if t in nhl_set:
        return t

    # Punctuation-stripped fallback
    import re
    stripped = re.sub(r'[^A-Z0-9]', '', t)
    if stripped in alias_map:
        return alias_map[stripped]

    # Map common 2-letter forms to NHL codes
    two_letter_map = {
        'TB': 'TBL',
        'LA': 'LAK',
        'NJ': 'NJD',
        'SJ': 'SJS',
        'MTL': 'MTL',  # already 3, keep as-is
        'WSH': 'WSH',  # keep if already NHL format
    }
    if stripped in two_letter_map:
        return two_letter_map[stripped]

    # Last resort: return cleaned code
    return stripped or t


def normalize_attrs_for_source(source, attrs, mappings):
    """Return a shallow-copied attributes dict with standardized team and position.
    Applies team mapping based on source and hardcoded position mapping.
    """
    attrs = dict(attrs or {})
    attrs['position'] = normalize_position(attrs.get('position'))
    team = attrs.get('last_seen_teamAbbrev')
    attrs['last_seen_teamAbbrev'] = normalize_team_for_source(source, team, mappings)
    return attrs

# Load team mappings once at import time
TEAM_MAPPINGS = load_team_mappings('Team2TM.xlsx')

class PlayerRegistry:
    def __init__(self):
        self.players = {}  # master_id -> {source_ids: {}, canonical_name: "", attributes: {}}
        self.next_id = 1
    
    def register_player(self, source, source_id, name, attributes=None):
        """
        Register a player from a data source, generating or retrieving a master ID.
        
        Args:
            source: The data source name (e.g., 'nhl_api', 'hockey_reference')
            source_id: ID from the source (or None if unavailable)
            name: Player name from the source
            attributes: Dict of additional attributes (birth_date, position, etc.)
            
        Returns:
            (master_id, created): Tuple where created=True if a new master record was created
        """
        if attributes is None:
            attributes = {}
        
        # Do not use volatile team fields to influence matching
        volatile_keys = {'last_seen_teamAbbrev', 'last_seen_teamName'}
        
        # Check if player already exists by source_id (only if source_id is provided)
        if source_id is not None:
            for master_id, player_data in self.players.items():
                if source in player_data['source_ids'] and player_data['source_ids'][source] == source_id:
                    # Update non-identifying rolling fields
                    for k, v in attributes.items():
                        if k in volatile_keys:
                            player_data['attributes'][k] = v
                    return master_id, False
        
        # If no match by ID, try fuzzy name matching with additional attributes
        normalized_name = self._normalize_name(name)

        # Team-aware, best-candidate fuzzy matching
        incoming_team = (attributes.get('last_seen_teamAbbrev') or '').strip().upper()
        incoming_pos = (attributes.get('position') or '').strip().upper()

        best_match_id = None
        best_score = -1.0
        best_player_ref = None

        for master_id, player_data in self.players.items():
            # Skip if we're sure this is a different player by having different IDs for same source
            if source in player_data['source_ids'] and player_data['source_ids'][source] != source_id and source_id is not None:
                continue

            existing_name = player_data['canonical_name']
            existing_attrs = player_data.get('attributes', {}) if isinstance(player_data.get('attributes'), dict) else {}
            existing_team = (existing_attrs.get('last_seen_teamAbbrev') or '').strip().upper()
            existing_pos = (existing_attrs.get('position') or '').strip().upper()

            # If both sides have non-empty teams and they differ, treat as a hard conflict
            if incoming_team and existing_team and incoming_team != existing_team:
                continue

            # Base similarity on name
            score = jellyfish.jaro_winkler_similarity(normalized_name, existing_name)

            # Team-aware boosts
            if incoming_team and existing_team and incoming_team == existing_team:
                score += 0.20

            # Position-aware scoring
            if incoming_pos and existing_pos:
                if incoming_pos == existing_pos:
                    score += 0.15
                else:
                    score -= 0.10  # Prefer the record with matching position when names are identical

            # Generic attribute boosts (exclude volatile team fields and position to avoid double-counting)
            for attr, value in attributes.items():
                if attr in volatile_keys or attr == 'position':
                    continue
                if attr in existing_attrs and existing_attrs[attr] == value:
                    score += 0.10

            if score > best_score:
                best_score = score
                best_match_id = master_id
                best_player_ref = player_data

        # Apply threshold and update best match if found
        if best_score >= 0.90 and best_match_id is not None:
            if source_id is not None:
                best_player_ref['source_ids'][source] = source_id
            # Update rolling team fields if provided
            for k, v in attributes.items():
                if k in volatile_keys:
                    best_player_ref['attributes'][k] = v
            return best_match_id, False
        
        # No match found, create new player
        master_id = self._generate_id()
        # Persist attributes including volatile fields as the initial snapshot
        src_ids = {}
        if source_id is not None:
            src_ids[source] = source_id
        self.players[master_id] = {
            'source_ids': src_ids,
            'canonical_name': normalized_name,
            'attributes': attributes
        }
        return master_id, True
    
    def _normalize_name(self, name):
        """Normalize player name by removing accents and standardizing format"""
        # Convert to lowercase
        name = name.lower()
        
        # Remove accents
        name = unicodedata.normalize('NFD', name).encode('ascii', 'ignore').decode('utf-8')
        
        # Handle common nicknames (could be expanded)
        nickname_map = {
            'alex': 'alexander',
            'mike': 'michael',
            # Add more as needed
        }
        
        # Split into parts and replace nicknames
        parts = name.split()
        if len(parts) > 0 and parts[0] in nickname_map:
            parts[0] = nickname_map[parts[0]]
        
        return ' '.join(parts)
    
    def _generate_id(self):
        """Generate a unique master ID for a new player"""
        master_id = self.next_id
        self.next_id += 1
        return master_id
    
    def get_player_sources(self, master_id):
        """Get all source IDs for a given master ID"""
        if master_id in self.players:
            return self.players[master_id]['source_ids']
        return None
    
    def get_canonical_name(self, master_id):
        """Get the canonical name for a player"""
        if master_id in self.players:
            return self.players[master_id]['canonical_name']
        return None
    
    def export_to_json(self, filename='player_registry.json'):
        """Export the entire player registry to JSON"""
        with open(filename, 'w') as f:
            json.dump(self.players, f, indent=2)
    
    def import_from_json(self, filename='player_registry.json'):
        """Import player registry from JSON"""
        try:
            with open(filename, 'r') as f:
                data = json.load(f)
                self.players = data
                # Ensure next_id is higher than any existing ID
                self.next_id = max([int(k) for k in self.players.keys()], default=0) + 1
        except FileNotFoundError:
            print(f"Registry file {filename} not found, starting with empty registry")

def get_data_from_url(team_abbrev):
    # Defensive sanitize in case input contains BOM/whitespace or mixed case
    if isinstance(team_abbrev, str):
        team_abbrev = team_abbrev.replace('\ufeff', '').strip().upper()
    url = f"https://api-web.nhle.com/v1/roster/{team_abbrev}/current"
    response = requests.get(url)
    print(url)

    if response.status_code != 200:
        print(f"Failed to get data for {team_abbrev}, status code: {response.status_code}")
        return None

    return response.json()


def read_team_abbrev_from_csv():
    team_abbreviations = []
    # Use utf-8-sig to automatically strip a UTF-8 BOM if present
    with open('Team2TM.csv', 'r', encoding='utf-8-sig', newline='') as file:
        csv_reader = csv.reader(file)
        for row in csv_reader:
            if not row:
                continue  # skip empty lines
            abbrev = str(row[0]).replace('\ufeff', '').strip().upper()
            if abbrev:
                team_abbreviations.append(abbrev)

    return team_abbreviations


def process_player_data(data, registry, source="nhl_api"):
    """Extract and register players from NHL API response data (robust to schema variants).
    Returns (processed_data, summary) where summary includes per-team and total new vs existing counts.
    """
    def _name_val(v):
        # Handle names that may be strings or objects like {"default": "John"}
        if isinstance(v, dict):
            return v.get("default") or v.get("first", "")
        return v or ""

    def _team_name(team):
        tn = team.get("teamName") or team.get("teamCommonName") or {}
        return _name_val(tn) if isinstance(tn, (dict, str)) else ""

    def _collect_players(team):
        # Different endpoints/layouts: some have consolidated 'roster', many have 'forwards','defensemen','goalies'
        buckets = []
        for key in ["roster", "forwards", "defensemen", "defense", "goalies", "skaters"]:
            val = team.get(key)
            if isinstance(val, list):
                buckets.append(val)
        players = []
        for bucket in buckets:
            for p in bucket:
                players.append(p)
        return players

    processed_data = []
    summary = {"totals": {"new": 0, "existing": 0}, "teams": []}

    for team_data in data:
        if not isinstance(team_data, dict):
            # Skip unexpected entries
            continue

        team_info = {
            "teamName": _team_name(team_data),
            "teamAbbrev": team_data.get("teamAbbrev", ""),
            "players": []
        }

        team_new = 0
        team_existing = 0

        players = _collect_players(team_data)
        processed_count = 0

        for player in players:
            # Extract robust ID field
            source_id = player.get("id") or player.get("playerId") or player.get("playerID")

            # Names may be nested objects
            first = _name_val(player.get("firstName"))
            last = _name_val(player.get("lastName"))
            name = (f"{first} {last}").strip()

            # Jersey/sweater number can be named differently
            jersey = player.get("jerseyNumber") or player.get("sweaterNumber") or ""

            position = player.get("positionCode") or player.get("position") or ""

            player_info = {
                "source_id": source_id,
                "name": name,
                "attributes": {
                    "position": position,
                    "jersey_number": jersey,
                    # Rolling fields to update in registry without affecting identity
                    "last_seen_teamAbbrev": team_info["teamAbbrev"],
                    "last_seen_teamName": team_info["teamName"],
                }
            }

            # Normalize attributes (position and team) to NHL standard before matching/registration
            player_info["attributes"] = normalize_attrs_for_source(source, player_info["attributes"], TEAM_MAPPINGS)

            # Register player and get master ID (only if we have at least a name or an ID)
            if source_id is not None or name:
                master_id, created = registry.register_player(
                    source,
                    player_info["source_id"],
                    player_info["name"],
                    player_info["attributes"]
                )
                player_info["master_id"] = master_id
                processed_count += 1
                if created:
                    team_new += 1
                else:
                    team_existing += 1

            team_info["players"].append(player_info)

        # Optional lightweight count for transparency
        if team_info["teamAbbrev"] or team_info["teamName"]:
            print(f"Processed {processed_count} players for {team_info['teamAbbrev'] or team_info['teamName']} (new: {team_new}, existing: {team_existing})")

        summary["teams"].append({
            "teamAbbrev": team_info["teamAbbrev"],
            "teamName": team_info["teamName"],
            "new": team_new,
            "existing": team_existing,
            "total": processed_count
        })
        summary["totals"]["new"] += team_new
        summary["totals"]["existing"] += team_existing

        processed_data.append(team_info)

    return processed_data, summary


def main():
    # Initialize player registry
    registry = PlayerRegistry()
    
    # Reset player registry file to empty
    try:
        if os.path.exists('player_registry.json'):
            with open('player_registry.json', 'w', encoding='utf-8') as f:
                f.write('{}')
    except Exception as e:
        print(f"Could not reset player_registry.json: {e}")
    
    # Try to load existing registry (now empty)
    registry.import_from_json()

    
    # Get data from NHL API
    team_abbreviations = read_team_abbrev_from_csv()
    data_list = []  # initialize an empty list to store all team data
    for team_abbrev in team_abbreviations:
        data = get_data_from_url(team_abbrev)
        if data is not None:
            data_list.append(data)  # append each team's data to the list

    # Optionally save the raw combined team data for transparency/debugging
    with open('all_teams_data.json', 'w', encoding='utf-8') as file:
        json.dump(data_list, file, indent=2)

    # Process player data and register players
    processed_data, summary = process_player_data(data_list, registry)

    # Save the processed data with assigned master IDs
    with open('all_teams_data_with_master_ids.json', 'w', encoding='utf-8') as file:
        json.dump(processed_data, file, indent=2)

    # Save a run summary for visibility (new vs existing)
    with open('run_summary.json', 'w', encoding='utf-8') as file:
        json.dump(summary, file, indent=2)

    # Fetch and process Natural Stat Trick player names (additional source)
    try:
        nst_processed, nst_summary = process_nst_player_names(registry)
        with open('nst_players.json', 'w', encoding='utf-8') as file:
            json.dump(nst_processed, file, indent=2)
        with open('nst_run_summary.json', 'w', encoding='utf-8') as file:
            json.dump(nst_summary, file, indent=2)
    except Exception as e:
        print(f"NST processing failed: {e}")

    # Fetch and process Hockey-Reference skaters (third source)
    try:
        hr_processed, hr_summary = process_hr_player_names(registry, season=2026)
        with open('hr_players.json', 'w', encoding='utf-8') as file:
            json.dump(hr_processed, file, indent=2)
        with open('hr_run_summary.json', 'w', encoding='utf-8') as file:
            json.dump(hr_summary, file, indent=2)
    except Exception as e:
        print(f"Hockey-Reference processing failed: {e}")

    # Fetch and process AG's Projections from Excel (fourth source)
    try:
        default_ag_path = "C:\\Users\\soluk\\OneDrive\\Documents\\FantasyNHL\\NatePts.xlsx"
        ag_processed, ag_summary = process_ag_player_names(
            registry,
            path=default_ag_path,
            sheet_name="NatePts",
            table_name="NatePts"
        )
        with open('ag_players.json', 'w', encoding='utf-8') as file:
            json.dump(ag_processed, file, indent=2)
        with open('ag_run_summary.json', 'w', encoding='utf-8') as file:
            json.dump(ag_summary, file, indent=2)
    except Exception as e:
        print(f"AG Projections processing failed: {e}")
    
    # Save the updated player registry
    registry.export_to_json()




# --- Natural Stat Trick integration: fetch and parse player names ---
from html.parser import HTMLParser


def fetch_nst_html(from_season="20252026", thru_season="20252026", stype="2"):
    """Fetch NaturalStatTrick player list page HTML (playerlist.php). Returns text or None on failure.
    Note: Parameters are kept for backward compatibility but not used by playerlist.php.
    """
    url = "https://www.naturalstattrick.com/playerlist.php"
    try:
        resp = requests.get(url, timeout=30)
        if resp.status_code == 200:
            return resp.text
        else:
            print(f"NST request failed: HTTP {resp.status_code}")
            return None
    except Exception as e:
        print(f"NST request error: {e}")
        return None


class _SimpleTableParser(HTMLParser):
    """Very small HTML table parser to gather headers and rows from the first table.
    Avoids external dependencies like BeautifulSoup.
    """
    def __init__(self):
        super().__init__()
        self.in_table = False
        self.in_tr = False
        self.in_th = False
        self.in_td = False
        self.headers = []
        self.rows = []
        self.current_row = []
        self.cell_buffer = []
        self.seen_first_table = False

    def handle_starttag(self, tag, attrs):
        if tag == "table" and not self.seen_first_table:
            self.in_table = True
        elif self.in_table and tag == "tr":
            self.in_tr = True
            self.current_row = []
        elif self.in_table and tag == "th":
            self.in_th = True
            self.cell_buffer = []
        elif self.in_table and tag == "td":
            self.in_td = True
            self.cell_buffer = []

    def handle_endtag(self, tag):
        if tag == "table" and self.in_table:
            self.in_table = False
            self.seen_first_table = True
        elif tag == "tr" and self.in_tr:
            self.in_tr = False
            # If we have header cells and no headers stored yet, treat as header row
            if self.headers:
                if self.current_row:
                    self.rows.append(self.current_row)
            else:
                if self.current_row:
                    self.headers = self.current_row
        elif tag == "th" and self.in_th:
            text = ("".join(self.cell_buffer)).strip()
            self.current_row.append(text)
            self.in_th = False
        elif tag == "td" and self.in_td:
            text = ("".join(self.cell_buffer)).strip()
            self.current_row.append(text)
            self.in_td = False

    def handle_data(self, data):
        if (self.in_th or self.in_td) and self.in_table and self.in_tr:
            self.cell_buffer.append(data)


def parse_nst_players(html_text):
    """Parse NST HTML table into a list of team dicts with players.
    Returns list shaped similarly to NHL team data processing for reuse.
    """
    if not html_text:
        return []
    parser = _SimpleTableParser()
    parser.feed(html_text)

    headers = [h.strip() for h in parser.headers]
    # Candidate column names
    player_keys = ["Player", "Name", "PLAYER"]
    team_keys = ["Team", "Tm", "TEAM", "TM"]
    pos_keys = ["Pos", "Position", "POSITION", "POS"]

    def find_idx(candidates):
        for c in candidates:
            if c in headers:
                return headers.index(c)
        # Try case-insensitive fallback
        lower = [h.lower() for h in headers]
        for c in candidates:
            if c.lower() in lower:
                return lower.index(c.lower())
        return -1

    pi = find_idx(player_keys)
    ti = find_idx(team_keys)
    poi = find_idx(pos_keys)
    if pi == -1:
        print("NST parse warning: Player column not found")
        return []
    # Team may be absent in some views; in that case group under Unknown

    teams = {}
    for row in parser.rows:
        if pi >= len(row):
            continue
        name = row[pi].strip()
        if not name:
            continue
        team = ""
        if ti != -1 and ti < len(row):
            team = row[ti].strip().upper()
        team = team or "UNKNOWN"
        pos = ""
        if poi != -1 and poi < len(row):
            pos = row[poi].strip().upper()
        if team not in teams:
            teams[team] = set()
        # Deduplicate only by (name, position) to retain distinct players with the same name
        teams[team].add((name, pos))

    processed = []
    for team_abbrev, name_pos_set in teams.items():
        team_info = {
            "teamName": "",
            "teamAbbrev": team_abbrev,
            "players": [{
                "source_id": "nst_playerlist",
                "name": n,
                "attributes": {
                    "position": p,
                    "jersey_number": "",
                    "last_seen_teamAbbrev": team_abbrev,
                    "last_seen_teamName": "",
                }
            } for (n, p) in sorted(name_pos_set)]
        }
        processed.append(team_info)
    return processed


def process_nst_player_names(registry, from_season="20252026", thru_season="20252026"):
    """Fetch and register player names from NaturalStatTrick.
    Returns (processed_data, summary) similar to process_player_data.
    """
    html = fetch_nst_html(from_season, thru_season)
    data = parse_nst_players(html)

    # Reuse registration flow with source name 'naturalstattrick'
    source = "naturalstattrick"
    processed_data = []
    summary = {"totals": {"new": 0, "existing": 0}, "teams": []}

    for team in data:
        raw_team_abbrev = team.get("teamAbbrev", "")
        team_abbrev = normalize_team_for_source(source, raw_team_abbrev, TEAM_MAPPINGS)
        team_name = team.get("teamName", "")
        team_new = 0
        team_existing = 0
        processed_count = 0
        out_team = {"teamName": team_name, "teamAbbrev": team_abbrev, "players": []}
        for p in team.get("players", []):
            name = p.get("name", "").strip()
            attrs = p.get("attributes", {})
            if not name:
                continue
            # normalize attrs before registering
            attrs = normalize_attrs_for_source(source, attrs, TEAM_MAPPINGS)
            master_id, created = registry.register_player(source, None, name, attrs)
            out_p = dict(p)
            # also reflect normalized attrs and team
            out_p["attributes"] = attrs
            out_p["master_id"] = master_id
            out_team["players"].append(out_p)
            processed_count += 1
            if created:
                team_new += 1
            else:
                team_existing += 1
        print(f"NST: Processed {processed_count} players for {team_abbrev or team_name} (new: {team_new}, existing: {team_existing})")
        summary["teams"].append({
            "teamAbbrev": team_abbrev,
            "teamName": team_name,
            "new": team_new,
            "existing": team_existing,
            "total": processed_count
        })
        summary["totals"]["new"] += team_new
        summary["totals"]["existing"] += team_existing
        processed_data.append(out_team)

    return processed_data, summary


# --- Hockey-Reference integration: fetch and parse skaters CSV ---

def fetch_hr_csv(season=2026):
    """Load Hockey-Reference skaters CSV for a given season from the local Downloads directory.
    Returns CSV text or None if the file is not found/readable.
    """
    try:
        # Default path: ~/Downloads/NHL_{season}_skaters.csv
        downloads_dir = os.path.join(os.path.expanduser('~'), 'Downloads')
        filename = f"NHL_{season}_skaters.csv"
        path = os.path.join(downloads_dir, filename)
        if not os.path.exists(path):
            print(f"HR local CSV not found at: {path}")
            return None
        with open(path, 'r', encoding='utf-8-sig', newline='') as f:
            text = f.read()
        print(f"HR: Loaded local CSV from {path} ({len(text)} bytes)")
        return text
    except Exception as e:
        print(f"HR local CSV read error: {e}")
        return None


def parse_hr_skaters_csv(csv_text):
    """Parse Hockey-Reference skaters CSV into list grouped by team, similar to NST/NHL structures.
    Handles files that include a two-level header by selecting the header row that contains 'Player'.
    """
    if not csv_text:
        print("HR: Empty CSV text provided to parser")
        return []
    import io
    import itertools

    sio = io.StringIO(csv_text)
    raw_reader = csv.reader(sio)
    rows = list(raw_reader)
    if not rows:
        print("HR: CSV has 0 rows")
        return []

    # Find the header row among the first 3 rows that contains 'Player'
    header_row_idx = None
    for i in range(min(3, len(rows))):
        cand = [ (c or '').strip() for c in rows[i] ]
        if any(h.lower() == 'player' for h in cand):
            header_row_idx = i
            header = cand
            break
    if header_row_idx is None:
        # Fallback: use the first row as header
        header_row_idx = 0
        header = [ (c or '').strip() for c in rows[0] ]
        print("HR parse notice: 'Player' not found in first 3 rows; using first row as header")
    else:
        if header_row_idx != 0:
            print(f"HR: Skipping {header_row_idx} pre-header row(s) before actual header")

    data_rows = rows[header_row_idx + 1:]

    # Build DictReader over the extracted header + data rows
    sio2 = io.StringIO()
    writer = csv.writer(sio2, lineterminator='\n')
    writer.writerow(header)
    writer.writerows(data_rows)
    sio2.seek(0)

    reader = csv.DictReader(sio2)

    # Build lowercase header map for flexible access
    def get(row, *keys):
        for k in keys:
            if k in row and row[k] is not None:
                return row[k]
        # try case-insensitive
        lower_map = { (kk or '').strip().lower(): (row[kk] if row[kk] is not None else '') for kk in row.keys() }
        for k in keys:
            v = lower_map.get((k or '').strip().lower())
            if v is not None:
                return v
        return ''

    teams = {}
    row_count = 0
    kept_count = 0
    for row in reader:
        row_count += 1
        name = (get(row, 'Player') or '').strip()
        if not name:
            continue
        # Skip aggregate rows like 'League Average'
        if name.strip().lower() == 'league average':
            continue
        team = (get(row, 'Tm', 'Team') or '').replace('\ufeff', '').strip().upper() or 'UNKNOWN'
        pos = (get(row, 'Pos', 'Position') or '').strip().upper()
        if team not in teams:
            teams[team] = set()
        teams[team].add((name, pos))
        kept_count += 1

    print(f"HR: Parsed {kept_count} player rows across {len(teams)} team buckets (from {row_count} data rows)")

    processed = []
    for team_abbrev, name_pos_set in teams.items():
        team_info = {
            "teamName": "",
            "teamAbbrev": team_abbrev,
            "players": [{
                "source_id": "hr_skaters_csv",
                "name": n,
                "attributes": {
                    "position": p,
                    "jersey_number": "",
                    "last_seen_teamAbbrev": team_abbrev,
                    "last_seen_teamName": "",
                }
            } for (n, p) in sorted(name_pos_set)]
        }
        processed.append(team_info)
    return processed


def fetch_hr_html(season=2026):
    """Fetch Hockey-Reference skaters HTML page for a given season (e.g., 2026). Returns HTML text or None."""
    url = f"https://www.hockey-reference.com/leagues/NHL_{season}_skaters.html"
    try:
        resp = requests.get(url, timeout=30)
        if resp.status_code == 200:
            return resp.text
        else:
            print(f"HR HTML request failed: HTTP {resp.status_code}")
            return None
    except Exception as e:
        print(f"HR HTML request error: {e}")
        return None


def parse_hr_skaters_html(html_text):
    """Parse Hockey-Reference skaters HTML table (#player_stats) into list grouped by team.
    Tries multiple robust strategies to extract the table even if wrapped in HTML comments or using
    alternate table IDs (e.g., skaters or player_stats), then uses _SimpleTableParser to read it.
    """
    if not html_text:
        return []
    import re

    table_html = None

    # 1) Look for table with id 'skaters' or 'player_stats' allowing single or double quotes
    id_pattern = r"<table[^>]*id=[\'\"](?:skaters|player_stats)[\'\"][^>]*>[\s\S]*?</table>"
    m = re.search(id_pattern, html_text, flags=re.IGNORECASE | re.DOTALL)

    # 2) If not found, scan commented blocks that may contain the table markup
    if not m:
        for cm in re.finditer(r"<!--[\s\S]*?-->", html_text, flags=re.IGNORECASE | re.DOTALL):
            block = cm.group(0)
            m2 = re.search(id_pattern, block, flags=re.IGNORECASE | re.DOTALL)
            if m2:
                m = m2
                break

    # 3) Fallback: find the container div and search inside it
    if not m:
        div_m = re.search(r"<div[^>]*id=[\'\"]all_skaters[\'\"][^>]*>([\s\S]*?)</div>", html_text, flags=re.IGNORECASE | re.DOTALL)
        if div_m:
            inside = div_m.group(1)
            m = re.search(id_pattern, inside, flags=re.IGNORECASE | re.DOTALL)

    if not m:
        print("HR parse warning: skaters table not found")
        return []

    table_html = m.group(0)

    parser = _SimpleTableParser()
    parser.feed(table_html)

    headers = [h.strip() for h in parser.headers]
    if not headers:
        print("HR parse warning: no headers parsed from skaters table")
        return []

    # Candidate column names
    player_keys = ["Player", "Name", "PLAYER"]
    team_keys = ["Team", "Tm", "TEAM", "TM"]
    pos_keys = ["Pos", "Position", "POSITION", "POS"]

    def find_idx(candidates):
        for c in candidates:
            if c in headers:
                return headers.index(c)
        lower = [h.lower() for h in headers]
        for c in candidates:
            if c.lower() in lower:
                return lower.index(c.lower())
        return -1

    pi = find_idx(player_keys)
    ti = find_idx(team_keys)
    poi = find_idx(pos_keys)
    if pi == -1:
        print("HR parse warning: Player column not found")
        return []

    teams = {}
    for row in parser.rows:
        if pi >= len(row):
            continue
        name = (row[pi] or "").strip()
        if not name:
            continue
        team = ""
        if ti != -1 and ti < len(row):
            team = (row[ti] or "").strip().upper()
        team = team or "UNKNOWN"
        pos = ""
        if poi != -1 and poi < len(row):
            pos = (row[poi] or "").strip().upper()
        if team not in teams:
            teams[team] = set()
        teams[team].add((name, pos))

    processed = []
    for team_abbrev, name_pos_set in teams.items():
        team_info = {
            "teamName": "",
            "teamAbbrev": team_abbrev,
            "players": [{
                "source_id": "hr_skaters_html",
                "name": n,
                "attributes": {
                    "position": p,
                    "jersey_number": "",
                    "last_seen_teamAbbrev": team_abbrev,
                    "last_seen_teamName": "",
                }
            } for (n, p) in sorted(name_pos_set)]
        }
        processed.append(team_info)
    return processed


def process_hr_player_names(registry, season=2026):
    """Load and register player names from a local Hockey-Reference skaters CSV for a given season.
    Returns (processed_data, summary) similar to process_player_data.
    """
    print(f"HR: Starting Hockey-Reference CSV processing for season {season}")
    csv_text = fetch_hr_csv(season)
    if not csv_text:
        print("HR: No CSV content to process")
    data = parse_hr_skaters_csv(csv_text)

    source = "hockey_reference"
    processed_data = []
    summary = {"totals": {"new": 0, "existing": 0}, "teams": []}

    if not data:
        print("HR: No teams parsed from CSV")

    for team in data:
        raw_team_abbrev = team.get("teamAbbrev", "")
        team_abbrev = normalize_team_for_source(source, raw_team_abbrev, TEAM_MAPPINGS)
        team_name = team.get("teamName", "")
        team_new = 0
        team_existing = 0
        processed_count = 0
        out_team = {"teamName": team_name, "teamAbbrev": team_abbrev, "players": []}
        for p in team.get("players", []):
            name = p.get("name", "").strip()
            attrs = p.get("attributes", {})
            if not name:
                continue
            attrs = normalize_attrs_for_source(source, attrs, TEAM_MAPPINGS)
            master_id, created = registry.register_player(source, None, name, attrs)
            out_p = dict(p)
            out_p["attributes"] = attrs
            out_p["master_id"] = master_id
            out_team["players"].append(out_p)
            processed_count += 1
            if created:
                team_new += 1
            else:
                team_existing += 1
        print(f"HR: Processed {processed_count} players for {team_abbrev or team_name} (new: {team_new}, existing: {team_existing})")
        summary["teams"].append({
            "teamAbbrev": team_abbrev,
            "teamName": team_name,
            "new": team_new,
            "existing": team_existing,
            "total": processed_count
        })
        summary["totals"]["new"] += team_new
        summary["totals"]["existing"] += team_existing
        processed_data.append(out_team)

    return processed_data, summary


def parse_ag_pts_xlsx(path, sheet_name="NatePts", table_name="NatePts"):
    """Parse AG's Projections Excel table into team-grouped structure.
    - path: absolute path to the Excel file
    - sheet_name: worksheet name that contains the table
    - table_name: Excel table name to read (preferred). If not found, falls back to using the sheet's used range.
    Returns list of {teamName, teamAbbrev, players: [ {source_id, name, attributes{position, jersey_number, last_seen_teamAbbrev, last_seen_teamName}} ]}
    """
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError:
        print("AG: openpyxl is not installed; cannot read Excel file.")
        return []
    try:
        if not path or not os.path.exists(path):
            print(f"AG: Excel file not found at: {path}")
            return []
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        # Determine cell range: prefer named table
        ref = None
        try:
            # openpyxl stores tables in ws.tables dict
            if hasattr(ws, 'tables') and table_name in getattr(ws, 'tables'):
                table_obj = ws.tables[table_name]
                ref = table_obj.ref
        except Exception:
            ref = None

        # Read headers and rows
        headers = []
        rows = []
        if ref:
            min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(ref)
            for r in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
                rows.append([ (str(c) if c is not None else '').strip() for c in r ])
        else:
            # Fallback: use the first contiguous block as data
            for r in ws.iter_rows(values_only=True):
                # convert to list of strings
                rows.append([ (str(c) if c is not None else '').strip() for c in r ])

        if not rows:
            print("AG: No rows found in worksheet")
            return []

        # Find header row containing 'Player' or similar
        header_row_idx = None
        for i in range(min(5, len(rows))):
            cand = [ (c or '').strip() for c in rows[i] ]
            lower = [x.lower() for x in cand]
            if any(x in lower for x in ('player', 'name', 'skater')):
                header_row_idx = i
                headers = cand
                break
        if header_row_idx is None:
            headers = [ (c or '').strip() for c in rows[0] ]
            header_row_idx = 0
            print("AG: Could not locate header with 'Player'; using first row as header")

        data_rows = rows[header_row_idx + 1:]

        # Helper to access fields case-insensitively
        lower_index = { (h or '').strip().lower(): idx for idx, h in enumerate(headers) }
        def get(row, *keys):
            for k in keys:
                i = lower_index.get((k or '').strip().lower())
                if i is not None and i < len(row):
                    return row[i]
            return ''

        # Build team buckets
        teams = {}
        kept = 0
        for row in data_rows:
            name = (get(row, 'Player', 'Name') or '').strip()
            if not name:
                continue
            team = (get(row, 'Team', 'Tm', 'TeamAbbrev') or '').replace('\ufeff', '').strip().upper() or 'UNKNOWN'
            pos = (get(row, 'Pos', 'Position') or '').strip().upper()
            if team not in teams:
                teams[team] = set()
            teams[team].add((name, pos))
            kept += 1
        print(f"AG: Parsed {kept} player rows across {len(teams)} team buckets")

        processed = []
        for team_abbrev, name_pos_set in teams.items():
            team_info = {
                "teamName": "",
                "teamAbbrev": team_abbrev,
                "players": [{
                    "source_id": "ag_pts_xlsx",
                    "name": n,
                    "attributes": {
                        "position": p,
                        "jersey_number": "",
                        "last_seen_teamAbbrev": team_abbrev,
                        "last_seen_teamName": "",
                    }
                } for (n, p) in sorted(name_pos_set)]
            }
            processed.append(team_info)
        return processed
    except Exception as e:
        print(f"AG: Excel parse error: {e}")
        return []


def process_ag_player_names(registry, path, sheet_name="NatePts", table_name="NatePts"):
    """Load and register player names from AG's Projections Excel file.
    Returns (processed_data, summary) similar to other sources.
    """
    data = parse_ag_pts_xlsx(path, sheet_name=sheet_name, table_name=table_name)

    source = "ag_projections"
    processed_data = []
    summary = {"totals": {"new": 0, "existing": 0}, "teams": []}

    if not data:
        print("AG: No teams parsed from Excel")

    for team in data:
        raw_team_abbrev = team.get("teamAbbrev", "")
        team_abbrev = normalize_team_for_source(source, raw_team_abbrev, TEAM_MAPPINGS)
        team_name = team.get("teamName", "")
        team_new = 0
        team_existing = 0
        processed_count = 0
        out_team = {"teamName": team_name, "teamAbbrev": team_abbrev, "players": []}
        for p in team.get("players", []):
            name = p.get("name", "").strip()
            attrs = p.get("attributes", {})
            if not name:
                continue
            attrs = normalize_attrs_for_source(source, attrs, TEAM_MAPPINGS)
            master_id, created = registry.register_player(source, None, name, attrs)
            out_p = dict(p)
            out_p["attributes"] = attrs
            out_p["master_id"] = master_id
            out_team["players"].append(out_p)
            processed_count += 1
            if created:
                team_new += 1
            else:
                team_existing += 1
        print(f"AG: Processed {processed_count} players for {team_abbrev or team_name} (new: {team_new}, existing: {team_existing})")
        summary["teams"].append({
            "teamAbbrev": team_abbrev,
            "teamName": team_name,
            "new": team_new,
            "existing": team_existing,
            "total": processed_count
        })
        summary["totals"]["new"] += team_new
        summary["totals"]["existing"] += team_existing
        processed_data.append(out_team)

    return processed_data, summary


if __name__ == "__main__":
    main()
